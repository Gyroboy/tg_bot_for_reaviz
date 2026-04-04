from __future__ import annotations

import logging
import os
import random
import re
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
from dotenv import load_dotenv
from telegram import InlineKeyboardButton, InlineKeyboardMarkup, ReplyKeyboardMarkup, Update
from telegram.ext import (
    Application,
    CallbackQueryHandler,
    CommandHandler,
    ContextTypes,
    MessageHandler,
    filters,
)


logging.basicConfig(
    format="%(asctime)s %(levelname)s %(name)s: %(message)s",
    level=logging.INFO,
)
LOGGER = logging.getLogger(__name__)

START_BUTTON = "Начать тест"
CHOOSE_COUNT_BUTTON = "Выбрать количество вопросов"
STOP_TEST_BUTTON = "Остановить тест"

OPTION_LETTERS = ["а", "б", "в", "г", "д", "е", "ж", "з", "и"]
LETTER_TO_INDEX = {letter: index for index, letter in enumerate(OPTION_LETTERS)}
QUESTION_FILE = Path("anatomy_test.xlsx")
ENV_FILES = (Path(".env"), Path("key.env"))


@dataclass(slots=True)
class Question:
    question_id: int
    text: str
    options: list[str]
    correct_indexes: list[int]
    question_type: str
    matching_groups: list[list[int]] = field(default_factory=list)
    matching_labels: list[str] = field(default_factory=list)

    @property
    def has_multiple_answers(self) -> bool:
        return self.question_type == "multi"

    @property
    def is_matching(self) -> bool:
        return self.question_type == "matching"

    def correct_option_text(self) -> str:
        if self.is_matching:
            parts = []
            for label, group in zip(self.matching_labels, self.matching_groups):
                mapped = ", ".join(str(index + 1) for index in group)
                parts.append(f"{label}) {mapped}")
            return "\n".join(parts)
        lines = []
        for index in self.correct_indexes:
            letter = OPTION_LETTERS[index]
            lines.append(f"{letter}) {self.options[index]}")
        return "\n".join(lines)


@dataclass(slots=True)
class TestSession:
    questions: list[Question]
    awaiting_count: bool = False
    current_index: int = 0
    correct_answers: int = 0
    selected_indexes: set[int] = field(default_factory=set)
    matching_step: int = 0
    matching_answers: list[list[int]] = field(default_factory=list)

    @property
    def total_questions(self) -> int:
        return len(self.questions)

    @property
    def current_question(self) -> Question:
        return self.questions[self.current_index]


def normalize_spaces(text: str) -> str:
    return re.sub(r"\s+", " ", text or "").strip()


def strip_option_prefix(text: str) -> str:
    return normalize_spaces(re.sub(r"^\s*[а-яa-z]\)\s*", "", text, flags=re.IGNORECASE))


def normalize_answer_string(raw_value: str) -> str:
    return (
        str(raw_value)
        .lower()
        .replace(" ", "")
        .replace("a", "а")
        .replace("b", "б")
        .replace("c", "в")
        .replace("d", "г")
        .replace("e", "д")
        .replace("f", "е")
        .replace("g", "ж")
        .replace("h", "з")
        .replace("i", "и")
    )


def parse_correct_indexes(raw_value: str) -> list[int]:
    indexes = []
    seen = set()
    latin_to_cyrillic = {"a": "а", "b": "б", "c": "в", "d": "г", "e": "д", "f": "е", "g": "ж", "h": "з", "i": "и"}
    digits = re.findall(r"\d", str(raw_value))
    if digits:
        for digit in digits:
            index = int(digit) - 1
            if 0 <= index < len(OPTION_LETTERS) and index not in seen:
                indexes.append(index)
                seen.add(index)
        if indexes:
            return indexes

    found_letters = re.findall(r"[а-яa-z]", str(raw_value).lower())

    for letter in found_letters:
        letter = latin_to_cyrillic.get(letter, letter)
        index = LETTER_TO_INDEX.get(letter)
        if index is not None and index not in seen:
            indexes.append(index)
            seen.add(index)
    return indexes


def parse_matching_groups(raw_value: str, options_count: int) -> tuple[list[str], list[list[int]]]:
    normalized = normalize_answer_string(raw_value)
    groups_with_letters = re.findall(r"([абвгдежзи])(\d+)", normalized)
    if groups_with_letters:
        labels = []
        groups = []
        for label, digits in groups_with_letters:
            labels.append(label)
            groups.append([int(digit) - 1 for digit in digits if digit.isdigit()])
        return labels, groups

    digits = [int(digit) - 1 for digit in normalized if digit.isdigit()]
    labels = OPTION_LETTERS[: min(len(digits), options_count)]
    groups = [[digit] for digit in digits]
    return labels, groups


def detect_question_type(raw_value: str) -> str:
    normalized = normalize_answer_string(raw_value)
    has_digits = any(char.isdigit() for char in normalized)
    if has_digits:
        return "matching"

    correct_indexes = parse_correct_indexes(raw_value)
    if len(correct_indexes) > 1:
        return "multi"
    return "single"


def load_questions(xlsx_path: Path) -> list[Question]:
    workbook = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    questions: list[Question] = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue

        question_text = normalize_spaces(str(row[0]))
        options = [strip_option_prefix(str(value)) for value in row[1:-1] if value]
        raw_answer = str(row[-1])
        question_type = detect_question_type(raw_answer)
        correct_indexes = parse_correct_indexes(raw_answer)
        matching_labels: list[str] = []
        matching_groups: list[list[int]] = []
        if question_type == "matching":
            matching_labels, matching_groups = parse_matching_groups(raw_answer, len(options))

        question_number_match = re.match(r"^\s*(\d+)", question_text)
        question_id = int(question_number_match.group(1)) if question_number_match else len(questions) + 1

        is_valid = bool(question_text and options and (correct_indexes or matching_groups))
        if is_valid:
            questions.append(
                Question(
                    question_id=question_id,
                    text=question_text,
                    options=options,
                    correct_indexes=correct_indexes,
                    question_type=question_type,
                    matching_labels=matching_labels,
                    matching_groups=matching_groups,
                )
            )

    return questions


def main_menu_keyboard() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup([[START_BUTTON]], resize_keyboard=True)


def test_menu_keyboard() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        [[CHOOSE_COUNT_BUTTON], [STOP_TEST_BUTTON]],
        resize_keyboard=True,
    )


def format_options_text(question: Question) -> str:
    """Форматирует варианты ответа как пронумерованный список для вывода в тексте сообщения."""
    lines = []
    for index, option in enumerate(question.options):
        letter = OPTION_LETTERS[index]
        lines.append(f"{letter}) {option}")
    return "\n".join(lines)


def single_choice_keyboard(question: Question) -> InlineKeyboardMarkup:
    """Кнопки содержат только букву варианта — полный текст выводится в сообщении."""
    buttons = [
        [InlineKeyboardButton(OPTION_LETTERS[index], callback_data=f"answer:{index}")]
        for index in range(len(question.options))
    ]
    return InlineKeyboardMarkup(buttons)


def multiple_choice_keyboard(question: Question, selected: set[int]) -> InlineKeyboardMarkup:
    """Кнопки содержат букву и галочку — полный текст выводится в сообщении."""
    buttons = []
    for index in range(len(question.options)):
        letter = OPTION_LETTERS[index]
        mark = "☑" if index in selected else "☐"
        buttons.append([InlineKeyboardButton(f"{mark} {letter}", callback_data=f"toggle:{index}")])
    buttons.append([InlineKeyboardButton("✅ Ответить", callback_data="submit")])
    return InlineKeyboardMarkup(buttons)


def matching_keyboard(question: Question, selected: set[int], step: int) -> InlineKeyboardMarkup:
    """Кнопки содержат номер и галочку — полный текст выводится в сообщении."""
    buttons = []
    for index in range(len(question.options)):
        mark = "☑" if index in selected else "☐"
        buttons.append(
            [InlineKeyboardButton(f"{mark} {index + 1}", callback_data=f"match_toggle:{index}")]
        )
    current_label = question.matching_labels[step]
    buttons.append([InlineKeyboardButton(f"✅ Готово для пункта {current_label}", callback_data="match_next")])
    return InlineKeyboardMarkup(buttons)


def get_session(context: ContextTypes.DEFAULT_TYPE, chat_id: int) -> TestSession | None:
    return context.chat_data.get("session")


def reset_session(context: ContextTypes.DEFAULT_TYPE) -> None:
    context.chat_data["session"] = None


async def show_test_menu(update: Update, text: str) -> None:
    if update.message:
        await update.message.reply_text(text, reply_markup=test_menu_keyboard())


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    reset_session(context)
    await update.message.reply_text(
        "Привет! Я помогу прорешивать тесты по анатомии.",
        reply_markup=main_menu_keyboard(),
    )


async def start_test_menu(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    session = TestSession(questions=[], awaiting_count=False)
    context.chat_data["session"] = session
    await show_test_menu(
        update,
        "Выберите действие: укажите количество вопросов для теста или остановите текущий тест.",
    )


async def ask_for_question_count(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    session = get_session(context, update.effective_chat.id)
    if session is None:
        session = TestSession(questions=[], awaiting_count=True)
        context.chat_data["session"] = session
    session.awaiting_count = True

    total_available = context.bot_data["total_questions"]
    await update.message.reply_text(
        f"Введите количество вопросов от 1 до {total_available}.",
        reply_markup=test_menu_keyboard(),
    )


async def stop_test(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    reset_session(context)
    await update.message.reply_text(
        "Тест остановлен. Когда захотите продолжить, нажмите «Начать тест».",
        reply_markup=main_menu_keyboard(),
    )


async def start_random_test(update: Update, context: ContextTypes.DEFAULT_TYPE, count: int) -> None:
    all_questions: list[Question] = context.bot_data["questions"]
    selected_questions = random.sample(all_questions, count)
    session = TestSession(questions=selected_questions)
    context.chat_data["session"] = session

    await update.message.reply_text(
        f"Отлично, начинаем тест из {count} вопросов.",
        reply_markup=test_menu_keyboard(),
    )
    await send_current_question(update.effective_chat.id, context)


async def send_current_question(chat_id: int, context: ContextTypes.DEFAULT_TYPE) -> None:
    session = get_session(context, chat_id)
    if session is None or session.current_index >= session.total_questions:
        return

    question = session.current_question
    options_text = format_options_text(question)
    header = f"Вопрос {session.current_index + 1} из {session.total_questions}\n\n{question.text}\n\n{options_text}"

    if question.is_matching:
        labels_text = ", ".join(question.matching_labels)
        prompt = (
            "\n\nЭто задание на соответствие."
            f"\nВыбирайте ответы по порядку для пунктов: {labels_text}."
            f"\nСейчас выберите вариант(ы) для пункта {question.matching_labels[session.matching_step]}."
            "\n\nНажимайте кнопки с номерами вариантов ниже."
        )
        keyboard = matching_keyboard(question, session.selected_indexes, session.matching_step)
    elif question.has_multiple_answers:
        prompt = "\n\nВ этом вопросе может быть несколько правильных ответов.\nВыберите все подходящие буквы и нажмите «Ответить»."
        keyboard = multiple_choice_keyboard(question, session.selected_indexes)
    else:
        prompt = "\n\nВыберите букву правильного варианта."
        keyboard = single_choice_keyboard(question)

    await context.bot.send_message(chat_id=chat_id, text=header + prompt, reply_markup=keyboard)


async def finish_test(chat_id: int, context: ContextTypes.DEFAULT_TYPE) -> None:
    session = get_session(context, chat_id)
    if session is None:
        return

    await context.bot.send_message(
        chat_id=chat_id,
        text=(
            f"Тест завершён.\n"
            f"Правильных ответов: {session.correct_answers} из {session.total_questions}."
        ),
        reply_markup=test_menu_keyboard(),
    )
    session.awaiting_count = False


async def move_to_next_question(chat_id: int, context: ContextTypes.DEFAULT_TYPE) -> None:
    session = get_session(context, chat_id)
    if session is None:
        return

    session.current_index += 1
    session.selected_indexes.clear()
    session.matching_step = 0
    session.matching_answers.clear()

    if session.current_index >= session.total_questions:
        await finish_test(chat_id, context)
        return

    await send_current_question(chat_id, context)


async def handle_count_input(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    session = get_session(context, update.effective_chat.id)
    if session is None or not session.awaiting_count:
        await update.message.reply_text(
            "Нажмите «Начать тест», чтобы открыть меню.",
            reply_markup=main_menu_keyboard(),
        )
        return

    text = normalize_spaces(update.message.text)
    if not text.isdigit():
        await update.message.reply_text("Введите число, например: 20")
        return

    count = int(text)
    total_available = context.bot_data["total_questions"]
    if not 1 <= count <= total_available:
        await update.message.reply_text(f"Введите число от 1 до {total_available}.")
        return

    await start_random_test(update, context, count)


def evaluate_answer(question: Question, selected_indexes: set[int]) -> tuple[bool, str]:
    correct_set = set(question.correct_indexes)
    is_correct = selected_indexes == correct_set
    correct_text = question.correct_option_text()

    if is_correct:
        return True, f"✅ Верно!\n\nПравильный ответ:\n{correct_text}"
    return False, f"❌ Неверно.\n\nПравильный ответ:\n{correct_text}"


def evaluate_matching_answer(question: Question, selected_groups: list[list[int]]) -> tuple[bool, str]:
    normalized_selected = [sorted(group) for group in selected_groups]
    normalized_expected = [sorted(group) for group in question.matching_groups]
    is_correct = normalized_selected == normalized_expected
    correct_text = question.correct_option_text()

    if is_correct:
        return True, f"✅ Верно! Соответствие указано правильно.\n\nПравильный ответ:\n{correct_text}"
    return False, f"❌ Неверно.\n\nПравильное соответствие:\n{correct_text}"


async def handle_single_answer(query, context: ContextTypes.DEFAULT_TYPE, option_index: int) -> None:
    session = get_session(context, query.message.chat_id)
    if session is None:
        await query.answer("Сначала начните тест.")
        return

    question = session.current_question
    is_correct, response_text = evaluate_answer(question, {option_index})
    if is_correct:
        session.correct_answers += 1

    await query.answer()
    await query.message.reply_text(response_text)
    await move_to_next_question(query.message.chat_id, context)


async def handle_multiple_answer_toggle(query, context: ContextTypes.DEFAULT_TYPE, option_index: int) -> None:
    session = get_session(context, query.message.chat_id)
    if session is None:
        await query.answer("Сначала начните тест.")
        return

    if option_index in session.selected_indexes:
        session.selected_indexes.remove(option_index)
    else:
        session.selected_indexes.add(option_index)

    question = session.current_question
    await query.answer("Ответ обновлён")
    await query.edit_message_reply_markup(
        reply_markup=multiple_choice_keyboard(question, session.selected_indexes)
    )


async def handle_multiple_answer_submit(query, context: ContextTypes.DEFAULT_TYPE) -> None:
    session = get_session(context, query.message.chat_id)
    if session is None:
        await query.answer("Сначала начните тест.")
        return

    if not session.selected_indexes:
        await query.answer("Сначала выберите хотя бы один вариант.")
        return

    question = session.current_question
    is_correct, response_text = evaluate_answer(question, session.selected_indexes)
    if is_correct:
        session.correct_answers += 1

    await query.answer()
    await query.message.reply_text(response_text)
    await move_to_next_question(query.message.chat_id, context)


async def handle_matching_toggle(query, context: ContextTypes.DEFAULT_TYPE, option_index: int) -> None:
    session = get_session(context, query.message.chat_id)
    if session is None:
        await query.answer("Сначала начните тест.")
        return

    if option_index in session.selected_indexes:
        session.selected_indexes.remove(option_index)
    else:
        session.selected_indexes.add(option_index)

    question = session.current_question
    await query.answer("Выбор обновлён")
    await query.edit_message_reply_markup(
        reply_markup=matching_keyboard(question, session.selected_indexes, session.matching_step)
    )


async def handle_matching_next(query, context: ContextTypes.DEFAULT_TYPE) -> None:
    session = get_session(context, query.message.chat_id)
    if session is None:
        await query.answer("Сначала начните тест.")
        return

    if not session.selected_indexes:
        await query.answer("Сначала выберите хотя бы один вариант.")
        return

    question = session.current_question
    session.matching_answers.append(sorted(session.selected_indexes))
    session.selected_indexes.clear()
    session.matching_step += 1

    if session.matching_step >= len(question.matching_labels):
        is_correct, response_text = evaluate_matching_answer(question, session.matching_answers)
        if is_correct:
            session.correct_answers += 1
        await query.answer()
        await query.message.reply_text(response_text)
        await move_to_next_question(query.message.chat_id, context)
        return

    await query.answer("Переходим к следующему пункту")
    await query.message.reply_text(
        f"Теперь выберите вариант(ы) для пункта {question.matching_labels[session.matching_step]}.",
        reply_markup=matching_keyboard(question, session.selected_indexes, session.matching_step),
    )


async def handle_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    await query.answer()

    data = query.data or ""
    if data.startswith("answer:"):
        await handle_single_answer(query, context, int(data.split(":", 1)[1]))
        return
    if data.startswith("toggle:"):
        await handle_multiple_answer_toggle(query, context, int(data.split(":", 1)[1]))
        return
    if data.startswith("match_toggle:"):
        await handle_matching_toggle(query, context, int(data.split(":", 1)[1]))
        return
    if data == "match_next":
        await handle_matching_next(query, context)
        return
    if data == "submit":
        await handle_multiple_answer_submit(query, context)


async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    text = normalize_spaces(update.message.text)

    if text == START_BUTTON:
        await start_test_menu(update, context)
        return
    if text == CHOOSE_COUNT_BUTTON:
        await ask_for_question_count(update, context)
        return
    if text == STOP_TEST_BUTTON:
        await stop_test(update, context)
        return

    await handle_count_input(update, context)


def build_application() -> Application:
    for env_file in ENV_FILES:
        load_dotenv(env_file)
    token = os.getenv("TELEGRAM_BOT_TOKEN")
    if not token:
        raise RuntimeError(
            "Не найден TELEGRAM_BOT_TOKEN. Создайте файл .env рядом с bot.py "
            "и добавьте строку TELEGRAM_BOT_TOKEN=ваш_токен"
        )

    questions = load_questions(QUESTION_FILE)
    if not questions:
        raise RuntimeError(f"Не удалось загрузить вопросы из файла {QUESTION_FILE}.")

    application = Application.builder().token(token).build()
    application.bot_data["questions"] = questions
    application.bot_data["total_questions"] = len(questions)

    application.add_handler(CommandHandler("start", start))
    application.add_handler(CallbackQueryHandler(handle_callback))
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))
    return application


def main() -> None:
    application = build_application()
    LOGGER.info("Бот загружен. Вопросов: %s", application.bot_data["total_questions"])
    application.run_polling()


if __name__ == "__main__":
    main()
