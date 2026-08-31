from __future__ import annotations

import re
from pathlib import Path

import openpyxl

from reaviz_bot.models import Question
from reaviz_bot.question_parser import AnswerParser
from reaviz_bot.text_utils import normalize_spaces, strip_option_prefix


class ExcelQuestionRepository:
    def __init__(self, xlsx_path: Path, answer_parser: AnswerParser | None = None) -> None:
        self.xlsx_path = xlsx_path
        self.answer_parser = answer_parser or AnswerParser()

    def load_questions(self) -> list[Question]:
        workbook = openpyxl.load_workbook(self.xlsx_path, read_only=True, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        questions: list[Question] = []

        rows = sheet.iter_rows(min_row=1, values_only=True)
        header = next(rows, None)
        answer_column = self._find_answer_column(header)

        for row in rows:
            question = self._parse_row(row, len(questions) + 1, answer_column)
            if question is not None:
                questions.append(question)

        return questions

    @staticmethod
    def _find_answer_column(header: tuple[object, ...] | None) -> int | None:
        """Индекс столбца с правильным ответом по заголовку.

        В файле по анатомии это последний столбец, а в файле по гистологии
        после него идут ещё служебные столбцы, поэтому ориентируемся на текст
        заголовка, а не на позицию.
        """

        if not header:
            return None
        for index, value in enumerate(header):
            if value and normalize_spaces(str(value)).lower().startswith("правильный ответ"):
                return index
        return None

    def _parse_row(
        self,
        row: tuple[object, ...],
        fallback_id: int,
        answer_column: int | None,
    ) -> Question | None:
        if not row or not row[0]:
            return None

        if answer_column is not None and answer_column < len(row):
            option_values = row[1:answer_column]
            raw_answer = str(row[answer_column]) if row[answer_column] is not None else ""
        else:
            option_values = row[1:-1]
            raw_answer = str(row[-1])

        question_text = normalize_spaces(str(row[0]))
        options = [strip_option_prefix(str(value)) for value in option_values if value]
        question_type = self.answer_parser.detect_question_type(raw_answer)
        correct_indexes = [
            index
            for index in self.answer_parser.parse_correct_indexes(raw_answer)
            if index < len(options)
        ]
        matching_labels: list[str] = []
        matching_groups: list[list[int]] = []

        if question_type == "matching":
            matching_labels, matching_groups = self.answer_parser.parse_matching_groups(raw_answer, len(options))

        question_number_match = re.match(r"^\s*(\d+)", question_text)
        question_id = int(question_number_match.group(1)) if question_number_match else fallback_id

        if not question_text or not options or not (correct_indexes or matching_groups):
            return None

        return Question(
            question_id=question_id,
            text=question_text,
            options=options,
            correct_indexes=correct_indexes,
            question_type=question_type,
            matching_labels=matching_labels,
            matching_groups=matching_groups,
        )

